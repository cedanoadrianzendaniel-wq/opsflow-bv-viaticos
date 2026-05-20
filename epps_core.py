"""Generador del F-004 Registro de Entrega de EPP.

Carga templates/epp_template.xlsx (con 17 EPPs hardcoded en filas 12-28) y rellena:
- Header (division, proyecto, nombre, dni, posicion)
- Tabla de EPPs: solo las filas correspondientes a los items entregados (resto queda en blanco)
- Observaciones
- Footer (Bureau Veritas + cliente)

Tambien procesa Excel masivo (formato 'Entrega de Epps {CLIENTE}.xlsx'):
- Cada fila = 1 trabajador, columnas F-Q son EPPs (talla o cantidad)
"""
import io, os, re, unicodedata
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XlsxImage

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'templates', 'epp_template.xlsx')
LOGO_PATH = os.path.join(os.path.dirname(__file__), 'templates', 'logo_bureau_veritas.jpg')

# Lista oficial de 17 EPPs (orden de filas en el template: row 12 -> EPP_LIST[0], etc)
EPP_LIST = [
    'Casco',
    'Barbiquejo',
    'Camisa manga larga ignifugo',
    'Pantalos ignifugo',          # ojo: typo del template original 'Pantalos' (no Pantalon)
    'Casaca',
    'Lentes claros',
    'Lentes oscuros',
    'Protector auditivo tipo copa',
    'Chaleco anaranjado',
    'Polo M/L',
    'Botas cana alta',
    'Baston Treckin',
    'Capotin',
    'Botas de PVC',
    'Guantes de badana',
    'Camisa Oxford',
    'Pantalon Jean',
]
# Row 12 -> index 0, row 28 -> index 16
EPP_ROW_BASE = 12


def _norm(s: str) -> str:
    if not s: return ''
    s = unicodedata.normalize('NFD', str(s))
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = re.sub(r'[^a-z0-9 ]', ' ', s.lower())
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def find_epp_row(nombre_epp: str):
    """Busca el row del template para un EPP. Devuelve row int o None."""
    target = _norm(nombre_epp)
    if not target: return None
    # Match exacto normalizado
    for idx, name in enumerate(EPP_LIST):
        if _norm(name) == target:
            return EPP_ROW_BASE + idx
    # Match por tokens: todos los tokens del target deben aparecer
    target_tokens = set(target.split())
    if not target_tokens: return None
    for idx, name in enumerate(EPP_LIST):
        name_tokens = set(_norm(name).split())
        # target_tokens subset de name_tokens, o name_tokens subset de target_tokens
        if target_tokens.issubset(name_tokens) or name_tokens.issubset(target_tokens):
            return EPP_ROW_BASE + idx
    # Match parcial: al menos 1 token coincide y es palabra "clave"
    keywords = {
        'casco': 0, 'barbiquejo': 1, 'camisa': None, 'pantalon': None, 'pantalos': 3,
        'casaca': 4, 'lentes': None, 'protector': 7, 'auditivo': 7, 'chaleco': 8,
        'polo': 9, 'botas': None, 'baston': 11, 'treckin': 11, 'trekking': 11,
        'capotin': 12, 'guantes': 14, 'badana': 14, 'oxford': 15, 'jean': 16,
    }
    for tok in target_tokens:
        idx = keywords.get(tok)
        if idx is not None:
            return EPP_ROW_BASE + idx
    # Camisa: ignifugo vs oxford
    if 'camisa' in target_tokens:
        if 'oxford' in target_tokens:
            return EPP_ROW_BASE + 15
        return EPP_ROW_BASE + 2  # ignifugo (default)
    # Pantalon: ignifugo vs jean
    if 'pantalon' in target_tokens or 'pantalones' in target_tokens:
        if 'jean' in target_tokens or 'jeans' in target_tokens:
            return EPP_ROW_BASE + 16
        return EPP_ROW_BASE + 3
    # Lentes: claros vs oscuros
    if 'lentes' in target_tokens:
        if 'oscuro' in target_tokens or 'oscuros' in target_tokens:
            return EPP_ROW_BASE + 6
        return EPP_ROW_BASE + 5
    # Botas: cana alta / pvc
    if 'botas' in target_tokens:
        if 'pvc' in target_tokens:
            return EPP_ROW_BASE + 13
        return EPP_ROW_BASE + 10
    return None


def _format_date(v):
    """Acepta datetime / 'YYYY-MM-DD' / 'DD/MM/YYYY' / None. Devuelve datetime o None."""
    if not v: return None
    if isinstance(v, datetime): return v
    s = str(v).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _add_months(d: datetime, months: int) -> datetime:
    """Suma N meses a una fecha. Si el dia destino no existe (ej 31 feb), usa el ultimo dia del mes."""
    import calendar
    new_month = d.month + months
    new_year = d.year + (new_month - 1) // 12
    new_month = ((new_month - 1) % 12) + 1
    last_day = calendar.monthrange(new_year, new_month)[1]
    new_day = min(d.day, last_day)
    return d.replace(year=new_year, month=new_month, day=new_day)


def generate_epp_excel(trabajador: dict, items: list, cliente_data: dict = None, observaciones: str = None) -> bytes:
    """Genera el F-004 lleno y devuelve bytes del xlsx.

    trabajador: {nombre_completo, dni, puesto, cliente, division}
    items: [{nombre_epp, cantidad, fecha_entrega, fecha_cambio, talla, cantidad_pendiente, descripcion_cambio}, ...]
    cliente_data: opcional, info adicional del cliente (no usado en este formato basico)
    observaciones: texto libre para celda de observaciones
    """
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb['F-004']

    # Re-insertar logo BV (openpyxl pierde la referencia al guardar)
    _logo_err = None
    if os.path.exists(LOGO_PATH):
        ws._images = []  # limpiar referencias rotas
        try:
            img = XlsxImage(LOGO_PATH)
            img.width = 110
            img.height = 80
            img.anchor = 'A1'
            ws.add_image(img)
        except Exception as e:
            import traceback
            _logo_err = f'{type(e).__name__}: {e} | {traceback.format_exc()[:500]}'
            print(f'[EPP LOGO ERROR] {_logo_err}')

    # Header
    ws['B5'] = trabajador.get('division') or 'INDUSTRIA'
    ws['B6'] = trabajador.get('cliente') or trabajador.get('proyecto') or ''
    ws['B7'] = trabajador.get('nombre_completo') or ''
    ws['B8'] = str(trabajador.get('dni') or '')
    ws['B9'] = trabajador.get('puesto') or trabajador.get('posicion') or ''

    # EPPs
    no_match = []
    for item in items:
        row = find_epp_row(item.get('nombre_epp') or item.get('nombre') or '')
        if row is None:
            no_match.append(item.get('nombre_epp') or item.get('nombre') or '?')
            continue
        if item.get('cantidad') is not None:
            ws[f'B{row}'] = item['cantidad']
        fe = _format_date(item.get('fecha_entrega'))
        if fe:
            ws[f'C{row}'] = fe
            ws[f'C{row}'].number_format = 'dd/mm/yyyy'
        fc = _format_date(item.get('fecha_cambio'))
        if not fc and fe:
            # Default: fecha_cambio = fecha_entrega + 6 meses
            fc = _add_months(fe, 6)
        if fc:
            ws[f'D{row}'] = fc
            ws[f'D{row}'].number_format = 'dd/mm/yyyy'
        if item.get('talla'):
            ws[f'E{row}'] = item['talla']
        if item.get('cantidad_pendiente') is not None:
            ws[f'F{row}'] = item['cantidad_pendiente']
        if item.get('descripcion_cambio'):
            ws[f'G{row}'] = item['descripcion_cambio']

    # Observaciones: A29 es anchor de merge A29:I29 con label "OBSERVACIONES: ". Append valor.
    if observaciones:
        ws['A29'] = f'OBSERVACIONES: {observaciones}'

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), {'no_match_epps': no_match, 'matched_count': len(items) - len(no_match), 'logo_err': _logo_err}


# ============================================================
# Excel masivo "Entrega de Epps {CLIENTE}.xlsx"
# ============================================================

# Mapeo de columnas del Excel masivo -> EPP_LIST index (None = no aplicable)
# A=nombre, B=proyecto, C=cc, D=puesto, E=fecha entrega
# F=Ignifugo Pantalon (talla, cant default 2) -> Pantalos ignifugo (idx 3)
# G=Ignifugo Camisa (talla, cant 2) -> Camisa ignifugo (idx 2)
# H=Pantalon Jean (talla, cant 2) -> Pantalon Jean (idx 16)
# I=Camisa Oxford (talla, cant 2) -> Camisa Oxford (idx 15)
# J=Chaleco (talla, cant 1) -> Chaleco anaranjado (idx 8)
# K=Zapato (talla, cant 1) -> Botas cana alta (idx 10)
# L=Capotin (cant) -> Capotin (idx 12)
# M=Casco (cant) -> Casco (idx 0)
# N=Tapanuca (cant) -> Barbiquejo (idx 1)
# O=Lentes (cant) -> Lentes claros (idx 5)
# P=Guantes (cant) -> Guantes badana (idx 14)
# Q=Orejera (cant) -> Protector auditivo (idx 7)
MASIVO_MAP = {
    'F': {'idx': 3, 'value_type': 'talla', 'default_cantidad': 2},
    'G': {'idx': 2, 'value_type': 'talla', 'default_cantidad': 2},
    'H': {'idx': 16, 'value_type': 'talla', 'default_cantidad': 2},
    'I': {'idx': 15, 'value_type': 'talla', 'default_cantidad': 2},
    'J': {'idx': 8, 'value_type': 'talla', 'default_cantidad': 1},
    'K': {'idx': 10, 'value_type': 'talla', 'default_cantidad': 1},
    'L': {'idx': 12, 'value_type': 'cantidad', 'default_talla': None},
    'M': {'idx': 0, 'value_type': 'cantidad'},
    'N': {'idx': 1, 'value_type': 'cantidad'},
    'O': {'idx': 5, 'value_type': 'cantidad'},
    'P': {'idx': 14, 'value_type': 'cantidad'},
    'Q': {'idx': 7, 'value_type': 'cantidad'},
}

COL_LETTER = {1:'A',2:'B',3:'C',4:'D',5:'E',6:'F',7:'G',8:'H',9:'I',10:'J',11:'K',12:'L',13:'M',14:'N',15:'O',16:'P',17:'Q'}


def parse_excel_masivo(content: bytes) -> list:
    """Parsea el Excel 'Entrega de Epps {cliente}.xlsx' y devuelve lista de trabajadores con sus items.

    Output: [
      {
        'nombre': str, 'proyecto': str, 'cc': str, 'puesto': str, 'fecha_entrega': str,
        'items': [{'nombre_epp':..., 'cantidad':..., 'talla':..., 'fecha_entrega':...}]
      }, ...
    ]
    """
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Validar headers en R1
    headers_esperados = ['COLABORADORES', 'PROYECTO']
    h1 = (ws['A1'].value or '').upper()
    if 'COLABORADOR' not in h1:
        raise ValueError(f'Formato no valido. Headers esperados: COLABORADORES en A1, encontrado: {h1!r}')

    out = []
    for row in range(2, ws.max_row + 1):
        nombre = ws[f'A{row}'].value
        if not nombre or not str(nombre).strip():
            continue
        proyecto = ws[f'B{row}'].value or ''
        cc = ws[f'C{row}'].value or ''
        puesto = ws[f'D{row}'].value or ''
        fecha_e = ws[f'E{row}'].value
        if isinstance(fecha_e, datetime):
            fecha_str = fecha_e.strftime('%Y-%m-%d')
        elif fecha_e:
            fecha_str = str(fecha_e)
        else:
            fecha_str = None

        items = []
        for col_letter, mapping in MASIVO_MAP.items():
            val = ws[f'{col_letter}{row}'].value
            if val is None or str(val).strip() == '':
                continue
            epp_name = EPP_LIST[mapping['idx']]
            item = {'nombre_epp': epp_name, 'fecha_entrega': fecha_str}

            if mapping['value_type'] == 'talla':
                # val es talla, cantidad va al default
                item['talla'] = str(val).strip()
                item['cantidad'] = mapping.get('default_cantidad', 1)
            else:
                # val es cantidad
                try:
                    item['cantidad'] = int(float(val))
                except (ValueError, TypeError):
                    continue
                if item['cantidad'] == 0:
                    continue  # 0 = no entregado

            items.append(item)

        out.append({
            'nombre': str(nombre).strip().upper(),
            'proyecto': str(proyecto).strip().upper(),
            'cc': str(cc).strip(),
            'puesto': str(puesto).strip().upper(),
            'fecha_entrega': fecha_str,
            'items': items,
        })
    return out
