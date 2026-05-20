"""Logica del procesador de viaticos. Stateless.

Inputs: bytes del Excel detalle, cliente, mes_label, datos de personal_bv y clientes_bv.
Outputs: bytes del consolidado xlsx, bytes del macro xlsm, metadata.
"""
import io, json, os, openpyxl, zipfile, re, html
from datetime import datetime
from copy import copy
from calendar import monthrange

# Templates path (relative al modulo)
TEMPLATES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates')
TEMPLATE_CONSOLIDADO = os.path.join(TEMPLATES_DIR, 'consolidado_template.xlsx')
TEMPLATE_MACRO = os.path.join(TEMPLATES_DIR, 'macro_template.xlsm')

# === MATCHING ===
def normalize_name(s):
    return ''.join(c for c in (s or '').upper().strip() if c.isalpha() or c == ' ').strip()

def levenshtein(a, b):
    if len(a) < len(b): a, b = b, a
    if not b: return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a):
        cur = [i + 1]
        for j, cb in enumerate(b):
            cur.append(min(prev[j + 1] + 1, cur[j] + 1, prev[j] + (ca != cb)))
        prev = cur
    return prev[-1]

def fuzzy_token(token, target_tokens):
    tol = 0 if len(token) <= 3 else (1 if len(token) <= 7 else 2)
    return any(levenshtein(token, t) <= tol for t in target_tokens)

def match_worker(name_input, personal):
    tokens_in = [t for t in normalize_name(name_input).split() if len(t) >= 3]
    if not tokens_in: return None
    best, best_score = None, 0
    for p in personal:
        target = [t for t in normalize_name(p.get('nombre_completo','')).split() if len(t) >= 3]
        score, ok = 0, True
        for tk in tokens_in:
            if tk in target: score += 1.0
            elif fuzzy_token(tk, target): score += 0.7
            else: ok = False; break
        if ok and score > best_score:
            best, best_score = p, score
    return best

# === PARSE INPUT ===
def parse_input_xlsx(content_bytes):
    """Lee el Excel detalle (bytes) y extrae los trabajadores con categorias F-ADM-002."""
    wb = openpyxl.load_workbook(io.BytesIO(content_bytes), data_only=True)
    s = wb.active
    workers = []
    for r in range(4, s.max_row + 1):
        nombre = s.cell(r, 2).value
        if not nombre or 'TOTAL' in str(nombre).upper(): continue
        def num(c): return float(s.cell(r, c).value or 0)
        w = {
            'numero': s.cell(r, 1).value,
            'nombre_input': str(nombre).strip(),
            'sector': str(s.cell(r, 3).value or '').strip().upper(),
            'localidad': str(s.cell(r, 4).value or '').strip(),
            'dias_servicio': int(num(6)),
            'alquiler_equipos_input': num(7), 'otros_input': num(8), 'copias_input': num(9),
            'bono': num(10), 'alimentacion': num(11), 'hospedaje': num(12), 'agua': num(13),
            'lavanderia': num(14), 'lavado_camioneta': num(15), 'cochera': num(16),
            'combustible': num(17), 'movilizacion': num(18), 'total_input': num(19),
        }
        w['cat_A'] = w['alimentacion'] + w['hospedaje'] + w['agua']
        w['cat_B'] = w['alquiler_equipos_input']
        w['cat_C'] = w['combustible'] + w['movilizacion']
        w['cat_D'] = w['lavanderia'] + w['lavado_camioneta'] + w['cochera']
        w['cat_E'] = w['copias_input'] + w['otros_input']
        w['total'] = w['cat_A'] + w['cat_B'] + w['cat_C'] + w['cat_D'] + w['cat_E']
        workers.append(w)
    return workers

# === GENERAR CONSOLIDADO XLSX ===
def calc_dates_mes(mes_label):
    y, m = mes_label.split('-')
    y, m = int(y), int(m)
    return datetime(y, m, 1), datetime(y, m, monthrange(y, m)[1])

def generate_consolidado_xlsx(workers_with_personal, cliente_data, mes_label):
    """Devuelve bytes del consolidado xlsx."""
    template = openpyxl.load_workbook(TEMPLATE_CONSOLIDADO)
    template_sheet = template['TGP']
    personal_template = template['Personal']
    fecha_salida, fecha_retorno = calc_dates_mes(mes_label)
    out = openpyxl.Workbook(); out.remove(out.active)

    for idx, item in enumerate(workers_with_personal, 1):
        w, p = item['worker'], item['personal']
        if not p: continue
        sheet_name = f"{idx:02d}_{(p['nombre_completo'] or '')[:25]}".replace('/','_').replace(':','_')[:31]
        ns = out.create_sheet(title=sheet_name)
        for row in template_sheet.iter_rows():
            for cell in row:
                nc = ns.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    nc.font = copy(cell.font); nc.fill = copy(cell.fill)
                    nc.border = copy(cell.border); nc.alignment = copy(cell.alignment)
                    nc.number_format = cell.number_format
        for col_letter, dim in template_sheet.column_dimensions.items():
            ns.column_dimensions[col_letter].width = dim.width
        for row_num, dim in template_sheet.row_dimensions.items():
            ns.row_dimensions[row_num].height = dim.height
        for mc in template_sheet.merged_cells.ranges:
            ns.merge_cells(str(mc))
        # Marcar checkbox Viaticos (E1) y agregar logo BV en A1
        ns.cell(row=1, column=5, value='X')
        try:
            from openpyxl.drawing.image import Image as XlsxImage
            logo_path = os.path.join(TEMPLATES_DIR, 'logo_bureau_veritas.jpg')
            if os.path.exists(logo_path):
                img = XlsxImage(logo_path)
                img.width = 110
                img.height = 80
                img.anchor = 'A1'
                ns.add_image(img)
        except Exception:
            pass
        ns.cell(row=3, column=3, value=datetime.now().strftime('%Y-%m-%d'))
        ns.cell(row=3, column=6, value=cliente_data.get('numero_contrato',''))
        ns.cell(row=6, column=4, value=p['nombre_completo'])
        ns.cell(row=7, column=4, value=p.get('puesto',''))
        ns.cell(row=8, column=4, value=str(p.get('dni','')))
        ns.cell(row=9, column=4, value=cliente_data.get('centro_costo',''))
        ns.cell(row=10, column=4, value=p.get('banco',''))
        ns.cell(row=11, column=4, value=p.get('cuenta_cci',''))
        ns.cell(row=12, column=4, value=f"Viaticos del mes - {cliente_data.get('cliente','')} - {mes_label}")
        cs = ns.cell(row=13, column=4, value=fecha_salida); cs.number_format = 'yyyy-mm-dd'
        cr = ns.cell(row=13, column=7, value=fecha_retorno); cr.number_format = 'yyyy-mm-dd'
        for r_, val in [(16, w['cat_A']), (17, w['cat_B']), (18, w['cat_C']), (19, w['cat_D']), (20, w['cat_E'])]:
            ns.cell(row=r_, column=4, value=val); ns.cell(row=r_, column=7, value=val)
        ns.cell(row=21, column=7, value=w['total'])
        ns.cell(row=23, column=1,
                value=f"Gastos mensuales asignados al periodo {mes_label}. Sector: {w['sector']}. Localidad: {w['localidad']}. Días servicio: {w['dias_servicio']}.")

    pers = out.create_sheet(title='Personal')
    for col in range(1, 11):
        h = personal_template.cell(1, col)
        nc = pers.cell(1, col, h.value)
        if h.has_style: nc.font = copy(h.font)
    for idx, item in enumerate(workers_with_personal, 1):
        w, p = item['worker'], item['personal']
        if not p: continue
        row = idx + 1
        pers.cell(row, 1, idx); pers.cell(row, 2, 'A')
        pers.cell(row, 3, p.get('tipo_cuenta_abono','A'))
        pers.cell(row, 4, p.get('cuenta_cci',''))
        pers.cell(row, 5, p.get('banco',''))
        pers.cell(row, 6, str(p.get('dni','')))
        pers.cell(row, 7, p.get('nombre_completo',''))
        pers.cell(row, 8, p.get('tipo_moneda','S'))
        pers.cell(row, 9, w['total'])
        pers.cell(row, 10, p.get('puesto',''))

    buf = io.BytesIO()
    out.save(buf)
    return buf.getvalue()

# === GENERAR MACRO XLSM (zip+regex preserva 100%) ===
def xml_escape(s):
    return html.escape(str(s), quote=False).replace('"','&quot;')

def generate_macro_xlsm(workers_with_personal, cliente_label):
    """Devuelve bytes del macro xlsm con Sueldos_WIN llena."""
    with open(TEMPLATE_MACRO, 'rb') as f:
        macro_bytes = f.read()
    src_buf = io.BytesIO(macro_bytes)
    with zipfile.ZipFile(src_buf, 'r') as zin:
        names_order = zin.namelist()
        files = {n: zin.read(n) for n in names_order}

    wb_xml = files['xl/workbook.xml'].decode('utf-8')
    m = re.search(r'<sheet[^/]*?name="Sueldos_WIN"[^/]*?r:id="(rId\d+)"', wb_xml)
    if not m: raise RuntimeError('No encontre Sueldos_WIN')
    rid = m.group(1)
    rels_xml = files['xl/_rels/workbook.xml.rels'].decode('utf-8')
    sheet_target = None
    for m_rel in re.finditer(r'<Relationship\b[^>]*/?>', rels_xml):
        if f'Id="{rid}"' in m_rel.group(0):
            tm = re.search(r'Target="([^"]+)"', m_rel.group(0))
            if tm: sheet_target = tm.group(1); break
    if not sheet_target: raise RuntimeError(f'No target para {rid}')
    sheet_path = 'xl/' + sheet_target if not sheet_target.startswith('xl/') else sheet_target

    shared_path = 'xl/sharedStrings.xml'
    ss_xml = files[shared_path].decode('utf-8')
    si_matches = list(re.finditer(r'<si\b[^>]*>(.*?)</si>', ss_xml, re.DOTALL))
    str_index = {}
    for i, sm in enumerate(si_matches):
        t_m = re.search(r'<t[^>]*>(.*?)</t>', sm.group(1), re.DOTALL)
        if t_m:
            str_index[html.unescape(t_m.group(1))] = i
    next_idx = len(si_matches)
    new_strings_xml = []

    def get_str_idx(s):
        nonlocal next_idx
        if s in str_index: return str_index[s]
        idx = next_idx
        str_index[s] = idx
        space_attr = ' xml:space="preserve"' if s != s.strip() else ''
        new_strings_xml.append(f'<si><t{space_attr}>{xml_escape(s)}</t></si>')
        next_idx += 1
        return idx

    sheet_xml = files[sheet_path].decode('utf-8')
    new_rows = []
    row_n = 8
    for item in workers_with_personal:
        w, p = item['worker'], item['personal']
        if not p: continue
        if not p.get('cuenta_cci') or not p.get('banco'): continue
        tipo_doc = p.get('tipo_documento') or 'DNI'
        doc_str = str(p.get('dni',''))
        if tipo_doc == 'CE' and len(doc_str) < 9:
            doc_str = doc_str.zfill(9)
        cells = [
            f'<c r="B{row_n}" t="s"><v>{get_str_idx("CARNET EXTRANJERÍA" if tipo_doc == "CE" else "DNI")}</v></c>',
            f'<c r="C{row_n}" t="s"><v>{get_str_idx(doc_str)}</v></c>',
            f'<c r="D{row_n}" t="s"><v>{get_str_idx(p.get("nombre_completo",""))}</v></c>',
            f'<c r="E{row_n}" t="s"><v>{get_str_idx("CTA. INTERBANCARIA SOLES")}</v></c>',
            f'<c r="G{row_n}" t="s"><v>{get_str_idx(p.get("cuenta_cci",""))}</v></c>',
            f'<c r="H{row_n}"><v>{w["total"]}</v></c>',
            f'<c r="I{row_n}" t="s"><v>{get_str_idx(f"VIATICO {cliente_label}")}</v></c>',
        ]
        correo = p.get('correo_corporativo') or p.get('correo_personal') or ''
        if correo:
            cells.append(f'<c r="J{row_n}" t="s"><v>{get_str_idx(correo)}</v></c>')
        new_rows.append(f'<row r="{row_n}">{"".join(cells)}</row>')
        row_n += 1

    new_rows_by_n = {8 + i: r for i, r in enumerate(new_rows)}

    def row_replacer(m):
        r_attr = int(m.group(1))
        if r_attr < 8 or r_attr > 30: return m.group(0)
        return new_rows_by_n.get(r_attr, '')

    sheet_xml_new = re.sub(r'<row\s+r="(\d+)"[^>]*?>.*?</row>', row_replacer, sheet_xml, flags=re.DOTALL)
    has_any = any(f'<row r="{n}"' in sheet_xml_new for n in new_rows_by_n)
    if not has_any:
        sheet_xml_new = sheet_xml_new.replace('</sheetData>', f'{"".join(new_rows)}</sheetData>')

    files[sheet_path] = sheet_xml_new.encode('utf-8')

    if new_strings_xml:
        ss_xml_new = re.sub(r'count="\d+"', f'count="{next_idx}"', ss_xml, count=1)
        ss_xml_new = re.sub(r'uniqueCount="\d+"', f'uniqueCount="{next_idx}"', ss_xml_new, count=1)
        ss_xml_new = ss_xml_new.replace('</sst>', f'{"".join(new_strings_xml)}</sst>')
        files[shared_path] = ss_xml_new.encode('utf-8')

    out_buf = io.BytesIO()
    with zipfile.ZipFile(out_buf, 'w', zipfile.ZIP_DEFLATED) as zout:
        for n in names_order:
            zout.writestr(n, files[n])
    return out_buf.getvalue()

# === API ===
def process_viaticos(content_bytes, cliente_label, mes_label, personal_list, cliente_data):
    """Procesa el Excel de viaticos y devuelve los 2 outputs + metadata."""
    workers = parse_input_xlsx(content_bytes)
    workers_matched = []
    for w in workers:
        p = match_worker(w['nombre_input'], personal_list)
        workers_matched.append({'worker': w, 'personal': p})

    consolidado = generate_consolidado_xlsx(workers_matched, cliente_data, mes_label)
    macro = generate_macro_xlsm(workers_matched, cliente_label)

    matched = [x for x in workers_matched if x['personal']]
    no_match = [x['worker']['nombre_input'] for x in workers_matched if not x['personal']]
    sin_cci = [x['personal']['nombre_completo'] for x in matched if not x['personal'].get('cuenta_cci')]
    total = sum(x['worker']['total'] for x in matched)

    return {
        'consolidado_xlsx': consolidado,
        'macro_xlsm': macro,
        'metadata': {
            'cliente': cliente_label,
            'mes_label': mes_label,
            'total_trabajadores_input': len(workers),
            'matched': len(matched),
            'no_match': no_match,
            'sin_cci': sin_cci,
            'total_monto': total,
        }
    }
