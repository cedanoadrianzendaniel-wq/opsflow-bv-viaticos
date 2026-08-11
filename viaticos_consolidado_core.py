"""Procesador del archivo 'Viaticos Proyecto Varios_{mes}.xlsx' (formato consolidado final, multi-cliente).

Estructura del archivo:
- Sheet unico con N bloques agrupados por proyecto.
- Cada bloque arranca con fila header:
    B: nombre proyecto (PLUSPETROL MNTTO MALVINAS / TGP MEDIO AMBIENTE / ...)
    C='CC', D='PROYECTO', E='Cargos', F-K: columnas de gastos, L='TOTAL VIATICO'
- Filas de datos: 1 trabajador por fila con B=nombre, C=CC, D=proyecto, E=cargo, F-K montos, L=total.
- Subtotal: fila con solo L con la suma del bloque.

Reusa generate_consolidado_xlsx y una version multi-cliente de generate_macro_xlsm de viaticos_core.py.
"""
import io, re, unicodedata
from openpyxl import load_workbook
from viaticos_core import match_worker, generate_consolidado_xlsx, generate_macro_xlsm


# Columnas (normalizadas) -> categoria F-ADM-002
COLUMN_TO_CATEGORY = {
    'alimentacion': 'cat_A',
    'hospedaje': 'cat_A',
    'alimentacion y hospedaje': 'cat_A',
    'agua': 'cat_A',
    'alquiler de equipos': 'cat_B',
    'alquiler equipos': 'cat_B',
    'combustible': 'cat_C',
    'movilizacion': 'cat_C',
    'transporte': 'cat_C',
    'lavanderia': 'cat_D',
    'lavado camioneta': 'cat_D',
    'cochera': 'cat_D',
    'lavado y limpieza': 'cat_D',
    'otros': 'cat_E',
    'bono': 'cat_E',  # confirmado por Daniel: BONO va en E
    'copias': 'cat_E',
}


def _norm_col(s):
    if not s: return ''
    s = unicodedata.normalize('NFD', str(s))
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return re.sub(r'\s+', ' ', s.strip().lower())


def _project_to_cliente(project_name):
    """Mapea nombre del bloque -> codigo cliente NocoDB."""
    u = (project_name or '').upper()
    if 'PLUSPETROL' in u or 'PPC' in u: return 'PPC-PLUSPETROL'
    if 'TGP' in u: return 'TGP'
    if 'TDP' in u: return 'TDP'
    if 'APM' in u: return 'APM'
    if 'COGA' in u: return 'TGP'
    return None


def _parse_single_cliente_sheet(wb, filename_hint=None):
    """Formato 'Viaticos {CLIENTE}' con 1 hoja de datos, columnas:
    NOMBRES, DNI/CE, CARGO, PROYECTO, BANCO, N.CUENTA, CCI, ALIMENTACION, ALQUILER, BONOS, OTROS, TOTAL.
    Busca la hoja que tenga esos headers (puede llamarse 'Viaticos APM', 'Viaticos TDP', etc.)
    Tambien acepta formato 'simple' tipo TGP: 1 hoja cualquiera con N + NOMBRE SUPERVISOR + categorias, sin DNI/cargo/total.
    Devuelve list[worker] o None si no detecta el formato.
    """
    HEADER_KEY_CAT = {'alimentacion','hospedaje','agua','alquiler','combustible','movilidad','movilizacion','transporte','lavanderia','lavado','cochera','otros','bonos','copias'}
    for sname in wb.sheetnames:
        ws = wb[sname]
        if ws.max_row < 3 or ws.max_column < 4:
            continue
        for header_row in (1, 2, 3):
            cols = {}
            for ci in range(1, ws.max_column + 1):
                h = _norm_col(ws.cell(row=header_row, column=ci).value)
                if h: cols[h] = ci
            has_nombres = any('nombre' in k or 'supervisor' in k or k.startswith('proyecto') or 'trabajador' in k or 'empleado' in k or 'colaborador' in k for k in cols)
            # Necesita al menos 1 columna de categoria
            has_categoria = any(any(p in k for p in HEADER_KEY_CAT) for k in cols)
            if not (has_nombres and has_categoria):
                continue
            # Mapeo de columnas (nombre puede ser "NOMBRE", "NOMBRES" o "NOMBRE SUPERVISOR")
            col_nombre = next((v for k,v in cols.items() if 'nombre' in k or 'supervisor' in k or k.startswith('proyecto') or 'trabajador' in k or 'empleado' in k or 'colaborador' in k), None)
            col_dni = next((v for k,v in cols.items() if 'dni' in k or k == 'ce'), None)
            col_cargo = next((v for k,v in cols.items() if 'cargo' in k), None)
            col_proyecto = next((v for k,v in cols.items() if 'proyecto' in k), None)
            col_total = next((v for k,v in cols.items() if 'total' in k), None)
            # Dedup bono: excluir cols con TODOS zeros (placeholders) o preferir neta (%, neto, 2.5)
            bono_keys = [k for k in cols if 'bono' in k]
            excluded = set()
            if len(bono_keys) > 1:
                # Check which bono cols have any non-zero data in rows
                data_start = header_row + 1
                bono_with_data = []
                for bk in bono_keys:
                    ci = cols[bk]
                    has = any(
                        (ws.cell(row=r, column=ci).value is not None) and
                        (isinstance(ws.cell(row=r, column=ci).value, (int, float)) and float(ws.cell(row=r, column=ci).value) != 0)
                        for r in range(data_start, min(data_start+50, ws.max_row+1))
                    )
                    if has: bono_with_data.append(bk)
                    else: excluded.add(ci)  # excluye placeholder vacio
                # Si aun quedan 2+, preferir la NETA
                if len(bono_with_data) > 1:
                    neto_keys = [k for k in bono_with_data if '%' in k or 'neto' in k or '2.5' in k]
                    keep = neto_keys[0] if neto_keys else bono_with_data[-1]
                    for bk in bono_with_data:
                        if bk != keep: excluded.add(cols[bk])
            cat_cols = []
            for k, v in cols.items():
                if v in excluded: continue
                if any(p in k for p in ['alimentacion','hospedaje','agua']):
                    cat_cols.append((v, 'cat_A'))
                elif 'alquiler' in k:
                    cat_cols.append((v, 'cat_B'))
                elif any(p in k for p in ['combustible','movilizacion','transporte','movilidad']):
                    cat_cols.append((v, 'cat_C'))
                elif any(p in k for p in ['lavanderia','lavado','cochera']):
                    cat_cols.append((v, 'cat_D'))
                elif any(p in k for p in ['otros','bonos','bono','copias','reembolso']):
                    cat_cols.append((v, 'cat_E'))
            if not cat_cols:
                continue
            # Inferir cliente del nombre de la hoja primero, despues del filename
            sname_upper = sname.upper()
            cliente = None
            for c in ('TGP','TDP','APM','PPC','COGA'):
                if c in sname_upper:
                    cliente = 'PPC-PLUSPETROL' if c == 'PPC' else c
                    break
            if not cliente and filename_hint:
                fn_upper = str(filename_hint).upper()
                for c in ('TGP','TDP','APM','PPC','COGA','PLUSPETROL'):
                    if c in fn_upper:
                        cliente = 'PPC-PLUSPETROL' if c in ('PPC','PLUSPETROL') else c
                        break
            if not cliente:
                cliente = 'BV'
            # Iterar filas de datos
            workers_out = []
            for r in range(header_row + 1, ws.max_row + 1):
                nombre = ws.cell(row=r, column=col_nombre).value
                if not nombre or not str(nombre).strip():
                    continue
                # Skip si la primera celda dice 'TOTAL' o algo asi
                if any(x in str(nombre).upper() for x in ('TOTAL','SUBTOTAL','SUMA')):
                    continue
                cats = {'cat_A':0.0,'cat_B':0.0,'cat_C':0.0,'cat_D':0.0,'cat_E':0.0}
                comments = {'cat_A':[], 'cat_B':[], 'cat_C':[], 'cat_D':[], 'cat_E':[]}
                for ci, cat in cat_cols:
                    cell = ws.cell(row=r, column=ci)
                    v = cell.value
                    if v is not None:
                        try: cats[cat] += float(v)
                        except (ValueError, TypeError): pass
                    if cell.comment and cell.comment.text:
                        txt = str(cell.comment.text).strip()
                        if txt: comments[cat].append(txt)
                try:
                    total = float(ws.cell(row=r, column=col_total).value) if col_total else sum(cats.values())
                except (ValueError, TypeError):
                    total = sum(cats.values())
                if total <= 0:
                    continue
                comment_total = ''
                if col_total:
                    tc = ws.cell(row=r, column=col_total)
                    if tc.comment and tc.comment.text:
                        comment_total = str(tc.comment.text).strip()
                workers_out.append({
                    'nombre': str(nombre).strip(),
                    'cliente': cliente,
                    'cliente_block_name': sname,
                    'cargo': str(ws.cell(row=r, column=col_cargo).value or '').strip() if col_cargo else '',
                    'sector': '', 'localidad': '', 'dias_servicio': '',
                    **cats,
                    'comments_by_cat': {k: ' | '.join(v) for k, v in comments.items()},
                    'comment_total': comment_total,
                    'total': total,
                })
            if workers_out:
                return workers_out
    return None


def parse_consolidado_final(content: bytes, filename_hint=None):
    """Detecta y parsea ambos formatos:
    1) Multi-bloques agrupados por proyecto (Viaticos Proyecto Varios)
    2) Single cliente con hoja 'Viaticos {CLIENTE}' y columnas fijas (Viaticos APM Terminals)
    Devuelve list of workers en estructura comun.
    """
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[wb.sheetnames[0]]

    workers = []
    current_block_cliente = None
    current_block_name = None
    current_cols_map = {}  # col_idx -> cat_X

    for r in range(1, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        d = ws.cell(row=r, column=4).value
        e = ws.cell(row=r, column=5).value
        l = ws.cell(row=r, column=12).value

        # Detectar header de bloque: C='CC' AND D='PROYECTO' AND E contiene 'Cargo'
        is_header = (
            isinstance(c, str) and c.strip().upper() == 'CC'
            and isinstance(d, str) and d.strip().upper() == 'PROYECTO'
            and isinstance(e, str) and 'cargo' in e.strip().lower()
        )
        if is_header:
            current_block_name = str(b or '').strip()
            current_block_cliente = _project_to_cliente(current_block_name)
            current_cols_map = {}
            for ci in range(6, 12):  # F..K
                hname = _norm_col(ws.cell(row=r, column=ci).value)
                cat = COLUMN_TO_CATEGORY.get(hname)
                if cat:
                    current_cols_map[ci] = cat
            continue

        # Fila trabajador: B con nombre + C con CC numerico + L con total
        if (current_block_cliente
            and isinstance(b, str) and b.strip()
            and (isinstance(c, (int, float)) or (isinstance(c, str) and c.strip().replace('.','').isdigit()))):
            cats = {'cat_A': 0.0, 'cat_B': 0.0, 'cat_C': 0.0, 'cat_D': 0.0, 'cat_E': 0.0}
            comments = {'cat_A': [], 'cat_B': [], 'cat_C': [], 'cat_D': [], 'cat_E': []}
            for ci, cat in current_cols_map.items():
                cell = ws.cell(row=r, column=ci)
                v = cell.value
                if v is None or (isinstance(v, str) and not v.strip()):
                    pass
                else:
                    try:
                        cats[cat] += float(v)
                    except (ValueError, TypeError):
                        pass
                if cell.comment and cell.comment.text:
                    txt = str(cell.comment.text).strip()
                    if txt:
                        comments[cat].append(txt)
            try:
                total = float(l)
            except (ValueError, TypeError):
                total = sum(cats.values())
            # Comentario del total (col L) si existe
            total_cell = ws.cell(row=r, column=12)
            comment_total = (total_cell.comment.text.strip() if total_cell.comment and total_cell.comment.text else '')
            workers.append({
                'nombre': b.strip(),
                'cliente': current_block_cliente,
                'cliente_block_name': current_block_name,
                'cargo': (e or '').strip() if isinstance(e, str) else '',
                'sector': '',       # no disponible en este formato
                'localidad': '',    # no disponible
                'dias_servicio': '',# no disponible
                **cats,
                'comments_by_cat': {k: ' | '.join(v) for k, v in comments.items()},
                'comment_total': comment_total,
                'total': total,
            })

    # Si no encontro bloques multi-cliente, probar formato single-cliente
    if not workers:
        single = _parse_single_cliente_sheet(wb, filename_hint=filename_hint)
        if single:
            workers = single

    # Fallback IA: si los parsers tradicionales fallan, llamar Claude
    if not workers:
        try:
            workers = _claude_parse_workers(wb)
        except Exception as e:
            print(f'[CLAUDE PARSE ERROR] {e}')

    return workers


def _excel_to_text(wb, max_rows=60, max_cols=20):
    """Convierte cada hoja del workbook a texto tipo tabla para enviar a Claude."""
    out = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        if ws.max_row == 0: continue
        out.append(f'\n=== Sheet: "{sname}" ({ws.max_row}x{ws.max_column}) ===')
        rows = []
        for r in range(1, min(ws.max_row + 1, max_rows + 1)):
            cells = []
            for c in range(1, min(ws.max_column + 1, max_cols + 1)):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    s = str(v).replace('\n', ' ').replace('\t', ' ')
                    if len(s) > 50: s = s[:47] + '...'
                    cells.append(f'{chr(64+c)}{r}={s}')
            if cells:
                rows.append(' | '.join(cells))
        out.append('\n'.join(rows))
    return '\n'.join(out)


def _claude_parse_workers(wb):
    """Llama Claude para extraer trabajadores + categorias F-ADM-002 del archivo.
    Usa ANTHROPIC_API_KEY env var. Devuelve list[worker] o None.
    """
    import os, json as _json
    api_key = os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        print('[CLAUDE] ANTHROPIC_API_KEY no configurada — fallback IA deshabilitado')
        return None

    from anthropic import Anthropic
    client = Anthropic(api_key=api_key)

    text = _excel_to_text(wb)

    system_prompt = """Eres un parser de archivos de viaticos para Bureau Veritas Peru. Recibes un Excel y extraes:
1. Lista de trabajadores con sus categorias F-ADM-002:
   - cat_A: Hospedaje y Alimentacion (alimentacion + hospedaje + agua + alim/hosp)
   - cat_B: Alquiler de Equipos
   - cat_C: Transporte (combustible + movilizacion)
   - cat_D: Lavado y Limpieza (lavanderia + lavado camioneta + cochera)
   - cat_E: Otros (otros + bonos + copias)
2. Cliente al que pertenece cada trabajador. Valores validos:
   TGP, TDP, APM, PPC-PLUSPETROL, COGA
   Inferir del nombre de la hoja, del proyecto, o del bloque del trabajador.

Devuelve SOLO un JSON valido (sin markdown, sin texto extra) con esta estructura exacta:
{
  "workers": [
    {
      "nombre": "NOMBRE COMPLETO",
      "cliente": "APM",
      "cargo": "SUPERVISOR HSE",
      "cat_A": 1500.0,
      "cat_B": 100.0,
      "cat_C": 0.0,
      "cat_D": 0.0,
      "cat_E": 0.0,
      "total": 1600.0
    }
  ]
}

Reglas:
- Solo trabajadores con total > 0.
- Ignora filas de subtotal/total general.
- Si una celda esta vacia o tiene #N/A, tratarla como 0.
- Los nombres en MAYUSCULAS."""

    msg = client.messages.create(
        model='claude-haiku-4-5-20251001',
        max_tokens=4000,
        system=system_prompt,
        messages=[{'role':'user','content': f'Archivo:\n\n{text}\n\nExtrae los trabajadores.'}]
    )
    raw = msg.content[0].text.strip()
    # Strip markdown si Claude lo agrega
    raw = re.sub(r'^```(?:json)?\s*', '', raw, flags=re.IGNORECASE)
    raw = re.sub(r'```\s*$', '', raw)
    raw = raw.strip()
    data = _json.loads(raw)
    workers_raw = data.get('workers', [])
    out = []
    for w in workers_raw:
        out.append({
            'nombre': str(w.get('nombre','')).strip().upper(),
            'cliente': w.get('cliente'),
            'cliente_block_name': w.get('cliente',''),
            'cargo': w.get('cargo',''),
            'sector': '', 'localidad': '', 'dias_servicio': '',
            'cat_A': float(w.get('cat_A', 0) or 0),
            'cat_B': float(w.get('cat_B', 0) or 0),
            'cat_C': float(w.get('cat_C', 0) or 0),
            'cat_D': float(w.get('cat_D', 0) or 0),
            'cat_E': float(w.get('cat_E', 0) or 0),
            'total': float(w.get('total', 0) or 0),
        })
    print(f'[CLAUDE] Extrajo {len(out)} trabajadores')
    return out


def process_consolidado_final(content: bytes, personal_list, clientes_list, mes_label, filename_hint=None):
    """Pipeline completo: parsea + matchea + genera consolidado + macro.
    Devuelve {'consolidado_xlsx': bytes, 'macro_xlsm': bytes, 'metadata': dict}.
    """
    parsed = parse_consolidado_final(content, filename_hint=filename_hint)
    if not parsed:
        raise ValueError('No se detectaron trabajadores en el archivo. Revisa estructura de bloques.')

    # Mapear cliente codigo -> cliente_data row
    clientes_by_code = {c.get('cliente'): c for c in clientes_list}

    workers_with_personal = []
    no_match = []
    for w in parsed:
        match = match_worker(w['nombre'], personal_list)
        if not match:
            no_match.append(w['nombre'])
            continue
        workers_with_personal.append({'worker': w, 'personal': match})

    # Para generate_consolidado_xlsx que pide cliente_data unico,
    # usamos el cliente mas comun en los workers como default.
    from collections import Counter
    cliente_counts = Counter(item['worker']['cliente'] for item in workers_with_personal)
    cliente_principal = cliente_counts.most_common(1)[0][0] if cliente_counts else None
    cliente_data = clientes_by_code.get(cliente_principal) or {'cliente': 'MULTI', 'numero_contrato': '', 'centro_costo': ''}

    # Generar consolidado xlsx
    consolidado = generate_consolidado_xlsx(workers_with_personal, cliente_data, mes_label)

    # Generar macro xlsm multi-cliente:
    # `generate_macro_xlsm` solo acepta UN cliente_label para todos. Workaround:
    # ejecutar por bloques pequeños... NO se puede combinar facil.
    # Solucion: agregamos al worker un campo `_referencia` con `VIATICO {cliente_block}`
    # y modificamos en runtime el helper (monkey-patch via wrapper si hace falta).
    # Por ahora: usar cliente_principal label para toda la macro.
    cliente_label = cliente_principal or 'BV'
    macro = generate_macro_xlsm(workers_with_personal, cliente_label)

    sin_cci = [item['personal'].get('nombre_completo','') for item in workers_with_personal
               if not item['personal'].get('cuenta_cci')]

    # Totales por cliente
    totals_por_cliente = {}
    for item in workers_with_personal:
        c = item['worker']['cliente']
        totals_por_cliente[c] = totals_por_cliente.get(c, 0) + item['worker']['total']

    # Detalle por trabajador (para enviar viatico individual despues del deposito)
    workers_detail = []
    for item in workers_with_personal:
        w = item['worker']
        p = item['personal']
        cmt = w.get('comments_by_cat', {}) or {}
        workers_detail.append({
            'dni': p.get('dni',''),
            'nombre_completo': p.get('nombre_completo',''),
            'cliente': w.get('cliente',''),
            'cargo': p.get('puesto') or w.get('cargo',''),
            'correo_personal': p.get('correo_personal',''),
            'correo_corporativo': p.get('correo_corporativo',''),
            'telefono': p.get('telefono',''),
            'banco': p.get('banco',''),
            'cuenta_cci': p.get('cuenta_cci',''),
            'items': {
                'alimentacion_hospedaje': float(w.get('cat_A', 0) or 0),
                'alquiler_equipos': float(w.get('cat_B', 0) or 0),
                'transporte_movilizacion': float(w.get('cat_C', 0) or 0),
                'lavado_limpieza': float(w.get('cat_D', 0) or 0),
                'otros_bonos': float(w.get('cat_E', 0) or 0),
            },
            'comments': {
                'alimentacion_hospedaje': cmt.get('cat_A',''),
                'alquiler_equipos': cmt.get('cat_B',''),
                'transporte_movilizacion': cmt.get('cat_C',''),
                'lavado_limpieza': cmt.get('cat_D',''),
                'otros_bonos': cmt.get('cat_E',''),
            },
            'comment_total': w.get('comment_total','') or '',
            'total': float(w.get('total', 0) or 0),
        })

    metadata = {
        'mes_label': mes_label,
        'matched': len(workers_with_personal),
        'no_match': no_match,
        'sin_cci': sin_cci,
        'totals_por_cliente': totals_por_cliente,
        'total_monto': sum(totals_por_cliente.values()),
        'cliente_principal': cliente_principal,
        'workers_detail': workers_detail,
    }
    return {'consolidado_xlsx': consolidado, 'macro_xlsm': macro, 'metadata': metadata}
